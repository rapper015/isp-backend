"""
Import legacy customer/subscriber records from an Excel export of the
client's existing ISP admin panel (IconRadius) into this system's
Customer/Subscriber/Plan models.

Usage:
    python manage.py import_legacy_users path/to/export.xlsx [--dry-run] [--sheet "Sheet1"] [--report import_report.csv]

Expected column headers (case-insensitive; matches the source panel's
"Customise Columns" labels on its /users list — extra or missing columns
are fine, matching is by header name not position):

    CAF No, Status, Username, Password, Name, Package Name, Mobile, Email,
    IpAddress, MAC, Expiry Date, City, State, Door No, Billing Address,
    Installation Address, GSTIN, Ref No, Current Balance

A Plan with a matching `name` (case-insensitive) must already exist in
this system for a row's subscriber to be created — this command does not
invent plans, since pricing/speed/pool details must be set up correctly
first. Rows whose package doesn't match an existing Plan are skipped and
logged to the report, not silently dropped.

Columns the source panel has that this schema has no field for yet —
Wallet Credit, Node/Pop/Switch, FUP Limit, Latitude/Longitude, Franchise/
Branch attribution on Customer/Subscriber, Account Type, Conn. Status/
Outage (live telemetry, not stored data), Commitment Date, Last Renewal,
Nas Port Id, Package Price/Custom Price — are NOT imported. CAF No, GSTIN,
Ref No, Door No, and State are folded into Customer.notes since there's no
dedicated field for them either, so the source values aren't lost even
though they're not queryable columns here.

Always run with --dry-run first and review the report before importing
for real.
"""
from __future__ import annotations

import csv
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from common.dates import parse_dt
from common.passwords import hash_password
from customers.models import Customer
from customers.sequences import next_customer_code
from plans.models import Plan
from subscribers.models import Subscriber

UNMAPPED_COLUMNS = [
    "Id", "Conn. Status", "Outage", "Account Type", "Franchise Name",
    "Sub Package", "Last Renewal", "User Type", "FUP Limit", "Area",
    "Colony", "Building", "Node", "Pop", "Switch", "Date Added",
    "User Added", "Commitment Date", "Package Price", "Custom Price",
    "Wallet Credit", "Last Logoff", "Nas Port Id", "Latitude", "Longitude",
]

STATUS_MAP = {
    "active": (Customer.Status.ACTIVE, Subscriber.Status.ACTIVE),
    "expired": (Customer.Status.ACTIVE, Subscriber.Status.SUSPENDED),
    "in-active": (Customer.Status.INACTIVE, Subscriber.Status.INACTIVE),
    "inactive": (Customer.Status.INACTIVE, Subscriber.Status.INACTIVE),
    "terminated": (Customer.Status.TERMINATED, Subscriber.Status.TERMINATED),
}


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _decimal(value, default=Decimal("0")) -> Decimal:
    text = _clean(value)
    if not text:
        return default
    try:
        return Decimal(text)
    except InvalidOperation:
        return default


class Command(BaseCommand):
    help = "Import legacy customer/subscriber records from an Excel export into Customer/Subscriber."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the .xlsx export file")
        parser.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
        parser.add_argument("--dry-run", action="store_true", help="Validate and report without writing to the DB")
        parser.add_argument("--report", default="import_report.csv", help="Path to write the per-row skip/warning report")

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl is required: pip install openpyxl") from exc

        path = options["path"]
        dry_run = options["dry_run"]
        report_path = options["report"]

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"File not found: {path}") from exc

        ws = wb[options["sheet"]] if options["sheet"] else wb.active
        rows = ws.iter_rows(values_only=True)
        header = [_clean(h) for h in next(rows)]
        col_index = {name.lower(): i for i, name in enumerate(header) if name}

        def get(row, name):
            idx = col_index.get(name.lower())
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        report_rows: list[list] = []
        stats = {"customers_created": 0, "customers_updated": 0,
                  "subscribers_created": 0, "subscribers_updated": 0, "skipped": 0}

        for line_no, row in enumerate(rows, start=2):
            if row is None or all(v in (None, "") for v in row):
                continue
            self._import_row(line_no, row, get, dry_run, report_rows, stats)

        with open(report_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["row", "username", "level", "message"])
            writer.writerows(report_rows)

        self.stdout.write(self.style.WARNING(
            "Source columns with no field to import into (dropped, not stored anywhere): "
            + ", ".join(UNMAPPED_COLUMNS)
        ))
        prefix = "[DRY RUN] " if dry_run else ""
        self.stdout.write(self.style.SUCCESS(
            f"{prefix}customers: {stats['customers_created']} created / {stats['customers_updated']} updated | "
            f"subscribers: {stats['subscribers_created']} created / {stats['subscribers_updated']} updated | "
            f"rows skipped: {stats['skipped']} | report: {report_path}"
        ))

    def _import_row(self, line_no, row, get, dry_run, report_rows, stats):
        username = _clean(get(row, "Username"))
        if not username:
            report_rows.append([line_no, "", "SKIPPED", "missing Username"])
            stats["skipped"] += 1
            return

        package_name = _clean(get(row, "Package Name"))
        if not package_name:
            report_rows.append([line_no, username, "SKIPPED", "missing Package Name"])
            stats["skipped"] += 1
            return

        plan = Plan.objects.filter(name__iexact=package_name, deleted_at__isnull=True).first()
        if plan is None:
            report_rows.append([line_no, username, "SKIPPED", f"Plan '{package_name}' not found in Plan catalog"])
            stats["skipped"] += 1
            return

        name = _clean(get(row, "Name")) or username
        mobile = _clean(get(row, "Mobile"))
        email = _clean(get(row, "Email"))
        caf_no = _clean(get(row, "CAF No"))
        city = _clean(get(row, "City"))
        state = _clean(get(row, "State"))
        billing_addr = _clean(get(row, "Billing Address"))
        install_addr = _clean(get(row, "Installation Address")) or billing_addr
        gstin = _clean(get(row, "GSTIN"))
        ref_no = _clean(get(row, "Ref No"))
        door_no = _clean(get(row, "Door No"))
        mac = _clean(get(row, "MAC"))
        ip_address = _clean(get(row, "IpAddress"))
        balance = _decimal(get(row, "Current Balance"))
        expires_at = parse_dt(get(row, "Expiry Date"))
        password_plain = _clean(get(row, "Password"))

        notes = " | ".join(
            part for part in (
                f"CAF No: {caf_no}" if caf_no else "",
                f"GSTIN: {gstin}" if gstin else "",
                f"Ref No: {ref_no}" if ref_no else "",
                f"Door No: {door_no}" if door_no else "",
                f"State: {state}" if state else "",
            ) if part
        )

        customer_status, subscriber_status = STATUS_MAP.get(
            _clean(get(row, "Status")).lower(), (Customer.Status.INACTIVE, Subscriber.Status.INACTIVE)
        )

        with transaction.atomic():
            customer = None
            if caf_no:
                customer = Customer.objects.filter(customer_code=caf_no).first()
            if customer is None and mobile:
                customer = Customer.objects.filter(phone=mobile).first()

            if customer is None:
                customer = Customer(
                    customer_code=caf_no or next_customer_code(),
                    full_name=name,
                    phone=mobile,
                    email=email,
                    address=install_addr,
                    city=city,
                    status=customer_status,
                    notes=notes,
                )
                if not dry_run:
                    customer.save()
                stats["customers_created"] += 1
            else:
                customer.full_name = name or customer.full_name
                customer.phone = mobile or customer.phone
                customer.email = email or customer.email
                customer.address = install_addr or customer.address
                customer.city = city or customer.city
                customer.status = customer_status
                customer.notes = notes or customer.notes
                if not dry_run:
                    customer.save()
                stats["customers_updated"] += 1

            subscriber = Subscriber.objects.filter(username=username).first()
            if subscriber is None:
                subscriber = Subscriber(
                    subscriber_code=username,
                    customer=customer,
                    plan=plan,
                    username=username,
                    password_hash=hash_password(password_plain) if password_plain else "",
                    status=subscriber_status,
                    installation_address=install_addr,
                    current_balance=balance,
                    static_ip_address=ip_address,
                    mac_address=mac,
                    expires_at=expires_at,
                )
                if not dry_run:
                    subscriber.save()
                stats["subscribers_created"] += 1
            else:
                subscriber.customer = customer
                subscriber.plan = plan
                subscriber.status = subscriber_status
                subscriber.installation_address = install_addr or subscriber.installation_address
                subscriber.current_balance = balance
                subscriber.static_ip_address = ip_address or subscriber.static_ip_address
                subscriber.mac_address = mac or subscriber.mac_address
                subscriber.expires_at = expires_at or subscriber.expires_at
                if password_plain:
                    subscriber.password_hash = hash_password(password_plain)
                if not dry_run:
                    subscriber.save()
                stats["subscribers_updated"] += 1

        if not password_plain:
            report_rows.append([
                line_no, username, "WARNING",
                "no Password in source row; subscriber has no usable password_hash and needs a reset before RADIUS login works",
            ])
